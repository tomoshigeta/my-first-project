# -*- coding: utf-8 -*-
"""入力シート（data/input-sheet.xlsx）を my-assets.json に変換する。

  python3 tools/xlsx_to_json.py data/input-sheet.xlsx data/my-assets.json

ダッシュボード本体は外部依存ゼロのまま。これは JSON を書き起こすための
別ツールで、画面を動かすのに必要なものではない。

金額は入力欄からこちらで計算し直す（Excel の計算結果は当てにしない）。
  総投資額     = 自己資金 + 当初借入額
  累計返済額   = 当初借入額 − 残債
  固定資産税月 = round(年額 / 12)
これで data/README.md の検算ルールが必ず成り立つ。
"""
import json
import sys
from datetime import date, datetime

from openpyxl import load_workbook

CLASSES = ["株", "債券", "投資信託", "その他"]
REGION_KEY = {"国内": "domestic", "海外": "overseas"}


class InputError(Exception):
    pass


def is_example(v):
    """データではない行（空行・見本行・注記・合計行）は読み飛ばす。"""
    if v is None:
        return True
    s = str(v).strip()
    return (s == ""
            or s.startswith("（例）") or s.startswith("(例)")
            or s.startswith("※")
            or s.startswith("合計"))


def num(v, where, allow_blank=False):
    if v is None or (isinstance(v, str) and not v.strip()):
        if allow_blank:
            return 0
        raise InputError(f"{where} が空です")
    if isinstance(v, bool) or not isinstance(v, (int, float)):
        raise InputError(f"{where} が数字ではありません（{v!r}）")
    f = round(float(v), 4)
    return int(f) if f == int(f) else f   # 1.0 ではなく 1 と書く


def text(v, where, allow_blank=False):
    if v is None or not str(v).strip():
        if allow_blank:
            return ""
        raise InputError(f"{where} が空です")
    return str(v).strip()


def as_of(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return text(v, "基本情報の基準日")


def convert(path):
    wb = load_workbook(path, data_only=True)
    for name in ("基本情報", "不動産", "有価証券", "預貯金", "その他の借入"):
        if name not in wb.sheetnames:
            raise InputError(f"「{name}」タブが見つかりません")

    base = wb["基本情報"]
    rates = {}
    for r in range(10, 14):
        code = base.cell(row=r, column=1).value
        val = base.cell(row=r, column=2).value
        if code is None or not str(code).strip():
            continue
        rates[str(code).strip().upper()] = num(val, f"基本情報の {code} のレート")
    if rates.get("JPY") != 1:
        raise InputError("基本情報の JPY のレートは 1 でなければなりません")

    data = {
        "formatVersion": 1,
        "owner": {"name": text(base["B4"].value, "基本情報の氏名")},
        "asOf": as_of(base["B5"].value),
        "rates": rates,
        "realEstate": [],
        "securities": [],
        "deposits": [],
        "otherLoans": [],
    }

    # ---------- 不動産 ----------
    ws = wb["不動産"]
    for r in range(5, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if is_example(name):
            continue
        w = f"不動産 {r}行目（{str(name).strip()}）"
        own = num(ws.cell(row=r, column=6).value, f"{w} の自己資金")
        principal = num(ws.cell(row=r, column=7).value, f"{w} の当初借入額")
        balance = num(ws.cell(row=r, column=8).value, f"{w} の残債")
        if balance > principal:
            raise InputError(f"{w}: 残債が当初借入額より大きくなっています")
        tax_year = num(ws.cell(row=r, column=16).value, f"{w} の固定資産税(年額)", allow_blank=True)
        data["realEstate"].append({
            "name": str(name).strip(),
            "address": text(ws.cell(row=r, column=2).value, f"{w} の住所", allow_blank=True),
            "areaSqm": num(ws.cell(row=r, column=3).value, f"{w} の広さ", allow_blank=True),
            "floor": text(ws.cell(row=r, column=4).value, f"{w} の階数", allow_blank=True),
            "ageYears": num(ws.cell(row=r, column=5).value, f"{w} の築年数", allow_blank=True),
            "totalInvestment": own + principal,
            "ownFunds": own,
            "loan": {
                "lender": text(ws.cell(row=r, column=9).value, f"{w} の借入先", allow_blank=True),
                "principal": principal,
                "balance": balance,
                "cumulativeRepaid": principal - balance,
                "rate": num(ws.cell(row=r, column=10).value, f"{w} の金利", allow_blank=True),
                "termYears": num(ws.cell(row=r, column=11).value, f"{w} の借入期間", allow_blank=True),
                "monthlyRepayment": num(ws.cell(row=r, column=12).value, f"{w} の毎月の返済額", allow_blank=True),
            },
            "monthly": {
                "rentIncome": num(ws.cell(row=r, column=13).value, f"{w} の家賃収入", allow_blank=True),
                "managementFee": num(ws.cell(row=r, column=14).value, f"{w} の管理費", allow_blank=True),
                "repairReserve": num(ws.cell(row=r, column=15).value, f"{w} の修繕積立金", allow_blank=True),
                "propertyTaxMonthly": int(round(tax_year / 12)),
            },
        })

    # ---------- 有価証券（国内・海外 × 4種類 = 8行で固定） ----------
    ws = wb["有価証券"]
    seen = set()
    for r in range(6, 14):
        region = text(ws.cell(row=r, column=1).value, f"有価証券 {r}行目 の地域")
        cls = text(ws.cell(row=r, column=2).value, f"有価証券 {r}行目 の種類")
        if region not in REGION_KEY:
            raise InputError(f"有価証券 {r}行目: 地域は「国内」か「海外」です（{region}）")
        if cls not in CLASSES:
            raise InputError(f"有価証券 {r}行目: 種類は {' / '.join(CLASSES)} のどれかです（{cls}）")
        cur = text(ws.cell(row=r, column=3).value, f"有価証券 {r}行目 の通貨").upper()
        if cur not in rates:
            raise InputError(f"有価証券 {r}行目: 通貨「{cur}」のレートが基本情報にありません")
        row = {
            "region": REGION_KEY[region],
            "assetClass": cls,
            "currency": cur,
            "cost": num(ws.cell(row=r, column=4).value, f"有価証券 {r}行目 の取得原価", allow_blank=True),
            "marketValue": num(ws.cell(row=r, column=5).value, f"有価証券 {r}行目 の評価額", allow_blank=True),
        }
        note = ws.cell(row=r, column=6).value
        if note and str(note).strip():
            row["note"] = str(note).strip()
        seen.add((row["region"], cls))
        data["securities"].append(row)
    missing = {(r, c) for r in REGION_KEY.values() for c in CLASSES} - seen
    if missing:
        raise InputError(f"有価証券の8行がそろっていません: {sorted(missing)}")

    # ---------- 預貯金 ----------
    ws = wb["預貯金"]
    for r in range(5, ws.max_row + 1):
        bank = ws.cell(row=r, column=1).value
        if is_example(bank):
            continue
        w = f"預貯金 {r}行目（{str(bank).strip()}）"
        cur = text(ws.cell(row=r, column=2).value, f"{w} の通貨").upper()
        if cur not in rates:
            raise InputError(f"{w}: 通貨「{cur}」のレートが基本情報にありません")
        data["deposits"].append({
            "bank": str(bank).strip(),
            "currency": cur,
            "amount": num(ws.cell(row=r, column=3).value, f"{w} の残高"),
        })

    # ---------- その他の借入 ----------
    ws = wb["その他の借入"]
    for r in range(6, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if is_example(name):
            continue
        w = f"その他の借入 {r}行目（{str(name).strip()}）"
        principal = num(ws.cell(row=r, column=4).value, f"{w} の当初借入額")
        balance = num(ws.cell(row=r, column=5).value, f"{w} の残債")
        if balance > principal:
            raise InputError(f"{w}: 残債が当初借入額より大きくなっています")
        data["otherLoans"].append({
            "name": str(name).strip(),
            "type": text(ws.cell(row=r, column=2).value, f"{w} の種類", allow_blank=True),
            "lender": text(ws.cell(row=r, column=3).value, f"{w} の借入先", allow_blank=True),
            "principal": principal,
            "balance": balance,
            "rate": num(ws.cell(row=r, column=6).value, f"{w} の金利", allow_blank=True),
            "monthlyRepayment": num(ws.cell(row=r, column=7).value, f"{w} の毎月の返済額", allow_blank=True),
        })

    return data


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        return 2
    src = sys.argv[1]
    dst = sys.argv[2] if len(sys.argv) > 2 else "data/my-assets.json"
    try:
        data = convert(src)
    except InputError as e:
        print(f"入力シートに直すところがあります:\n  - {e}")
        return 1
    with open(dst, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    print(f"{dst} を書きました（不動産 {len(data['realEstate'])}件 / "
          f"預貯金 {len(data['deposits'])}件 / その他の借入 {len(data['otherLoans'])}件）")
    return 0


if __name__ == "__main__":
    sys.exit(main())
