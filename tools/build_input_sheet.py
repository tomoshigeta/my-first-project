# -*- coding: utf-8 -*-
"""data/input-sheet.xlsx（入力シート）を組み立てる。

このシートは JSON を手書きする代わりの下書きで、正はあくまで my-assets.json。
足し算で出せる数字（総投資額・累計返済額・固定資産税の月額）は Excel の数式に
持たせ、不変条件が入力の時点で必ず成り立つようにしている。

  python3 tools/build_input_sheet.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

F = "Arial"
TITLE  = Font(name=F, size=14, bold=True)
H2     = Font(name=F, size=11, bold=True)
BODY   = Font(name=F, size=10)
HEAD   = Font(name=F, size=10, bold=True, color="FFFFFF")
INPUT  = Font(name=F, size=10, color="0000FF")        # 青字 = 手で入れる
CALC   = Font(name=F, size=10, color="000000")        # 黒字 = 自動計算
EX     = Font(name=F, size=10, color="808080", italic=True)
MUTED  = Font(name=F, size=9,  color="595959")

FILL_HEAD = PatternFill("solid", fgColor="2F5597")
FILL_IN   = PatternFill("solid", fgColor="FFF2CC")    # 薄い黄 = 入力欄
FILL_CALC = PatternFill("solid", fgColor="F2F2F2")    # 薄い灰 = 自動計算
FILL_EX   = PatternFill("solid", fgColor="FAFAFA")
FILL_NG   = PatternFill("solid", fgColor="FFC7CE")

THIN = Side(style="thin", color="BFBFBF")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

YEN  = '#,##0;(#,##0);-'
PCT1 = '0.0%'
RATE = '0.00"%"'
NUM1 = '0.0'

wb = Workbook()

def head_row(ws, row, headers, widths, calc_from=None):
    """見出し行を敷く。calc_from 以降の列は自動計算の扱い。"""
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font, c.fill, c.border = HEAD, FILL_HEAD, BOX
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 34
    ws.freeze_panes = ws.cell(row=row + 1, column=1)

def style_block(ws, r0, r1, ncols, calc_cols, fmts, example_row=None):
    """データ行に色と罫線を敷く。calc_cols は 1 始まりの列番号の集合。"""
    for r in range(r0, r1 + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BOX
            if r == example_row:
                cell.font, cell.fill = EX, FILL_EX
            elif c in calc_cols:
                cell.font, cell.fill = CALC, FILL_CALC
            else:
                cell.font, cell.fill = INPUT, FILL_IN
            if c in fmts:
                cell.number_format = fmts[c]

def note(ws, row, text, col=1):
    c = ws.cell(row=row, column=col, value=text)
    c.font = MUTED
    return c

# ============================================================
# はじめに
# ============================================================
ws = wb.active
ws.title = "はじめに"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 11      # 色見本を置く狭い列
ws.column_dimensions["C"].width = 96      # 本文（B からはみ出して表示される）

def line(row, text, font=BODY):
    """本文は B に書く。C が空なので右へはみ出して1行で読める。"""
    c = ws.cell(row=row, column=2, value=text)
    c.font = font
    c.alignment = Alignment(vertical="center")
    return c

def legend(row, fill, label, font=None):
    sw = ws.cell(row=row, column=2, value=None)
    sw.fill, sw.border = fill, BOX
    if font:
        sw.value, sw.font = "（例）", font
    t = ws.cell(row=row, column=3, value=label)
    t.font = BODY
    t.alignment = Alignment(vertical="center")

line(2, "資産ダッシュボード 入力シート", TITLE)
line(4, "使い方は3ステップ", H2)
line(5, "1. 下のタブ（基本情報・不動産・有価証券・預貯金・その他の借入）を順に埋めます。")
line(6, "2. 「ファイル → 名前を変えて保存」で my-assets.xlsx として保存します（この白紙のひな形は残しておきます）。")
line(7, "3. このファイルを Claude とのチャットに送ります。my-assets.json に変換してお返しします。")
line(9, "セルの色の意味", H2)

legend(10, FILL_IN,   "薄い黄色　＝　あなたが入れる欄（文字は青）")
legend(11, FILL_CALC, "薄い灰色　＝　Excel が自動で計算する欄（さわらないでください）")
legend(12, FILL_EX,   "灰色の斜体　＝　書き方の見本。消しても残しても構いません（変換のとき読み飛ばします）", font=EX)

line(14, "入力するときの決まり", H2)
line(15, "・金額は「円」で入れます。1万円は 10000 です。カンマ（,）は自分で打たなくて構いません。")
line(16, "・「月」と書いてある欄は 1か月分の金額です。固定資産税だけは年額を入れてください（月額は自動で計算します）。")
line(17, "・数字は半角で入れます。全角の１２３は使えません。")
line(18, "・持っていない項目は 0 のままにします。行ごと消す必要はありません。")
line(19, "・行が足りなくなったら、一番下の行をコピーして貼り付けて増やしてください。")
line(21, "気をつけていただきたいこと", H2)
line(22, "このファイルには氏名と資産の全額が入ります。メールやクラウドに置いたままにせず、Mac の中で保管してください。")
line(23, "GitHub には上げません（data/ フォルダは白紙のひな形と説明書以外が除外してあります）。")
line(25, "「検算」タブに合計と、つじつまの合わない箇所の警告が出ます。送る前に一度ご覧ください。")

# ============================================================
# 基本情報
# ============================================================
ws = wb.create_sheet("基本情報")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 26
ws.column_dimensions["C"].width = 60

ws["A1"] = "基本情報"; ws["A1"].font = TITLE
ws["A3"] = "項目"; ws["B3"] = "入力"; ws["C3"] = "説明"
for col in "ABC":
    ws[f"{col}3"].font, ws[f"{col}3"].fill, ws[f"{col}3"].border = HEAD, FILL_HEAD, BOX

rows = [
    ("氏名",   "山田 太郎",   "画面の左上に出ます"),
    ("基準日", "2026-09-01",  "いつ時点の数字か。2026-09-01 の形で入れます"),
]
for i, (k, v, d) in enumerate(rows, start=4):
    ws.cell(row=i, column=1, value=k).font = BODY
    c = ws.cell(row=i, column=2, value=v)
    c.font, c.fill, c.border = INPUT, FILL_IN, BOX
    c.alignment = Alignment(horizontal="left")
    ws.cell(row=i, column=1).border = BOX
    ws.cell(row=i, column=3, value=d).font = MUTED
    ws.cell(row=i, column=3).border = BOX
ws["B5"].number_format = "@"   # 日付は文字として保つ

ws["A7"] = "為替レート"; ws["A7"].font = H2
note(ws, 8, "外貨を持っている場合だけ入れます。「その日のレート」を手で固定します（自動では取りません）。", col=1)
ws["A9"] = "通貨"; ws["B9"] = "1単位あたりの円"; ws["C9"] = "説明"
for col in "ABC":
    ws[f"{col}9"].font, ws[f"{col}9"].fill, ws[f"{col}9"].border = HEAD, FILL_HEAD, BOX

rate_rows = [("JPY", 1, "日本円。1 のまま変えないでください"),
             ("USD", 155.2, "米ドル。1ドルが何円か"),
             ("", None, "他の通貨があれば（例: EUR）"),
             ("", None, "")]
for i, (k, v, d) in enumerate(rate_rows, start=10):
    a = ws.cell(row=i, column=1, value=k); b = ws.cell(row=i, column=2, value=v)
    if i == 10:
        a.font = b.font = CALC; a.fill = b.fill = FILL_CALC
    else:
        a.font = b.font = INPUT; a.fill = b.fill = FILL_IN
    a.border = b.border = BOX
    b.number_format = NUM1
    ws.cell(row=i, column=3, value=d).font = MUTED
    ws.cell(row=i, column=3).border = BOX
RATE_TBL = "'基本情報'!$A$10:$B$13"

# ============================================================
# 不動産
# ============================================================
ws = wb.create_sheet("不動産")
ws.sheet_view.showGridLines = False
ws["A1"] = "不動産　（1行に1つの物件）"; ws["A1"].font = TITLE
note(ws, 2, "灰色の「総投資額」「累計返済額」「固定資産税(月)」「月次収支」は自動で計算されます。")

RE_HEAD = ["物件名", "住所", "広さ(㎡)", "階数", "築年数",
           "自己資金", "当初借入額", "残債", "借入先", "金利(年%)",
           "借入期間(当初)\n年", "借入期間(当初)\nか月", "毎月の返済額",
           "家賃収入(月)", "管理費(月)", "修繕積立金(月)", "固定資産税(年額)",
           "総投資額", "累計返済額", "固定資産税(月)", "月次収支", "借入期間 合計(か月)"]
RE_W = [16, 24, 9, 12, 8, 13, 13, 13, 14, 10, 11, 11, 13, 13, 12, 13, 15, 14, 13, 14, 13, 15]
head_row(ws, 4, RE_HEAD, RE_W)
RE_R0, RE_R1 = 5, 14                      # 5 = 見本、6..14 = 入力
ws.cell(row=RE_R0, column=1, value="（例）物件A")
for col, val in zip(range(2, 18),
                    ["東京都港区白金台", 45.5, "3 / 10 階", 15,
                     10000000, 40000000, 35500000, "A銀行", 1.8, 26, 11, 128000,
                     180000, 12000, 8000, 108000]):
    ws.cell(row=RE_R0, column=col, value=val)
for r in range(RE_R0, RE_R1 + 1):
    ws.cell(row=r, column=18, value=f"=F{r}+G{r}")                       # 総投資額
    ws.cell(row=r, column=19, value=f"=G{r}-H{r}")                       # 累計返済額
    ws.cell(row=r, column=20, value=f"=ROUND(Q{r}/12,0)")                # 固定資産税(月)
    ws.cell(row=r, column=21, value=f"=N{r}-M{r}-O{r}-P{r}-T{r}")        # 月次収支
    ws.cell(row=r, column=22, value=f"=K{r}*12+L{r}")                    # 借入期間 合計(か月)
RE_FMT = {3: NUM1, 6: YEN, 7: YEN, 8: YEN, 10: RATE, 13: YEN,
          14: YEN, 15: YEN, 16: YEN, 17: YEN, 18: YEN, 19: YEN, 20: YEN, 21: YEN}
style_block(ws, RE_R0, RE_R1, 22, {18, 19, 20, 21, 22}, RE_FMT, example_row=RE_R0)
for c in (18, 19, 20, 21, 22):
    ws.cell(row=RE_R0, column=c).font = EX; ws.cell(row=RE_R0, column=c).fill = FILL_EX
ws.cell(row=4, column=17).comment = Comment(
    "年額を入れてください。1年ぶんです。右の「固定資産税(月)」は 12 で割った平均月額で、自動計算です。", "入力シート", width=280, height=90)
ws.cell(row=4, column=11).comment = Comment(
    "契約したときの期間です（残りの期間ではありません）。26年11か月なら「年」に 26、「か月」に 11。"
    "ちょうど20年なら 20 と 0。返済回数しか分からないときは 12 で割った商と余りを入れてください。",
    "入力シート", width=300, height=110)
ws.cell(row=4, column=7).comment = Comment(
    "借りたときの金額です。「残債」は今あといくら残っているか。差額が「累計返済額」として自動で出ます。", "入力シート", width=280, height=90)
note(ws, RE_R1 + 2, "※ 行が足りないときは、最終行（14行目）をコピーして下に貼り付けてください。数式も一緒に増えます。")

# ============================================================
# 有価証券
# ============================================================
ws = wb.create_sheet("有価証券")
ws.sheet_view.showGridLines = False
ws["A1"] = "有価証券　（国内・海外 × 4種類 ＝ 8行で固定）"; ws["A1"].font = TITLE
note(ws, 2, "持っていない種類は 0 のままにします（行は消しません。入れ忘れに気づけるようにするためです）。")
note(ws, 3, "「その他」は暗号資産・金・銀などをまとめる欄です。金額はその通貨のままで入れます（円換算は自動）。")

SEC_HEAD = ["地域", "種類", "通貨", "取得原価", "評価額", "補足（任意）",
            "含み損益", "円換算した評価額", "円換算した取得原価"]
head_row(ws, 5, SEC_HEAD, [10, 12, 9, 15, 15, 22, 15, 17, 17])
SEC_R0 = 6
combos = [("国内", c, "JPY") for c in ("株", "債券", "投資信託", "その他")] + \
         [("海外", c, "USD") for c in ("株", "債券", "投資信託", "その他")]
sample = [(2000000, 5000000, ""), (2700000, 2500000, ""), (1300000, 1500000, ""),
          (800000, 1200000, "金地金"), (20000, 32000, ""), (18000, 16000, ""),
          (8000, 9500, ""), (5000, 11000, "暗号資産")]
for i, ((reg, cls, cur), (cost, mv, nt)) in enumerate(zip(combos, sample)):
    r = SEC_R0 + i
    ws.cell(row=r, column=1, value=reg).font = CALC
    ws.cell(row=r, column=2, value=cls).font = CALC
    ws.cell(row=r, column=3, value=cur)
    ws.cell(row=r, column=4, value=cost)
    ws.cell(row=r, column=5, value=mv)
    ws.cell(row=r, column=6, value=nt)
    ws.cell(row=r, column=7, value=f"=E{r}-D{r}")
    rate_of = f"IFERROR(INDEX('基本情報'!$B$10:$B$13,MATCH(C{r},'基本情報'!$A$10:$A$13,0)),0)"
    ws.cell(row=r, column=8, value=f"=ROUND(E{r}*{rate_of},0)")   # 円換算した評価額
    ws.cell(row=r, column=9, value=f"=ROUND(D{r}*{rate_of},0)")   # 円換算した取得原価
SEC_R1 = SEC_R0 + 7
style_block(ws, SEC_R0, SEC_R1, 9, {1, 2, 7, 8, 9}, {4: YEN, 5: YEN, 7: YEN, 8: YEN, 9: YEN})
tot = SEC_R1 + 1
ws.cell(row=tot, column=1, value="合計").font = H2
for col, f in ((4, f"=SUM(D{SEC_R0}:D{SEC_R1})"), (5, f"=SUM(E{SEC_R0}:E{SEC_R1})"),
               (7, f"=SUM(G{SEC_R0}:G{SEC_R1})"), (8, f"=SUM(H{SEC_R0}:H{SEC_R1})"),
               (9, f"=SUM(I{SEC_R0}:I{SEC_R1})")):
    c = ws.cell(row=tot, column=col, value=f)
    c.font, c.number_format, c.border = H2, YEN, BOX
note(ws, tot, "※ 通貨が違うので、この行の取得原価と評価額の合計は目安です）", col=6)
dv = DataValidation(type="list", formula1='"JPY,USD,EUR"', allow_blank=False)
ws.add_data_validation(dv); dv.add(f"C{SEC_R0}:C{SEC_R1}")
ws.cell(row=5, column=3).comment = Comment(
    "「基本情報」タブの為替レートに書いた通貨だけ使えます。", "入力シート", width=260, height=70)

# ============================================================
# 預貯金
# ============================================================
ws = wb.create_sheet("預貯金")
ws.sheet_view.showGridLines = False
ws["A1"] = "預貯金　（1行に1つの口座）"; ws["A1"].font = TITLE
note(ws, 2, "同じ銀行に複数口座があるときは、まとめて1行にしても、分けて書いても構いません。")
head_row(ws, 4, ["金融機関", "通貨", "残高", "円換算した残高"], [26, 10, 16, 17])
DEP_R0, DEP_R1 = 5, 19
ws.cell(row=DEP_R0, column=1, value="（例）A銀行")
ws.cell(row=DEP_R0, column=2, value="JPY")
ws.cell(row=DEP_R0, column=3, value=2000000)
for r in range(DEP_R0, DEP_R1 + 1):
    ws.cell(row=r, column=4, value=f"=IF(A{r}=\"\",\"\",ROUND(C{r}*IFERROR(INDEX('基本情報'!$B$10:$B$13,MATCH(B{r},'基本情報'!$A$10:$A$13,0)),0),0))")
style_block(ws, DEP_R0, DEP_R1, 4, {4}, {3: YEN, 4: YEN}, example_row=DEP_R0)
dv2 = DataValidation(type="list", formula1='"JPY,USD,EUR"', allow_blank=True)
ws.add_data_validation(dv2); dv2.add(f"B{DEP_R0}:B{DEP_R1}")
c = ws.cell(row=DEP_R1 + 1, column=1, value="合計（見本行は数えません）"); c.font = H2
c = ws.cell(row=DEP_R1 + 1, column=4, value=f"=SUM(D{DEP_R0 + 1}:D{DEP_R1})")
c.font, c.number_format, c.border = H2, YEN, BOX

# ============================================================
# 保険
# ============================================================
ws = wb.create_sheet("保険")
ws.sheet_view.showGridLines = False
ws["A1"] = "保険　（1行に1つの契約）"; ws["A1"].font = TITLE
note(ws, 2, "貯蓄性のある保険（終身・養老・個人年金など）を書きます。掛け捨ての医療保険などは書きません。")
note(ws, 3, "評価額には「解約返戻金」（今解約したら戻る額）を入れます。毎年届く「ご契約内容のお知らせ」か保険会社のマイページで分かります。")
note(ws, 4, "※ 死亡保険金額は参考です。今使えるお金ではないので、資産の合計には足しません。")
head_row(ws, 6, ["保険名", "保険会社", "通貨", "払込保険料 累計", "解約返戻金",
                 "死亡保険金額\n(参考・資産に含めず)", "評価日", "補足（任意）",
                 "含み損益", "円換算した解約返戻金", "円換算した払込保険料"],
         [26, 16, 9, 16, 16, 17, 13, 22, 15, 18, 18])
INS_R0, INS_R1 = 7, 14
for col, val in zip(range(1, 9),
                    ["（例）変額終身保険", "○○生命", "JPY", 3000000, 2600000,
                     10000000, "2026-03-31", "解約返戻金は年1回の通知の数字"]):
    ws.cell(row=INS_R0, column=col, value=val)
for r in range(INS_R0, INS_R1 + 1):
    rate_of = f"IFERROR(INDEX('基本情報'!$B$10:$B$13,MATCH(C{r},'基本情報'!$A$10:$A$13,0)),0)"
    ws.cell(row=r, column=9,  value=f'=IF(A{r}="","",E{r}-D{r})')
    ws.cell(row=r, column=10, value=f'=IF(A{r}="","",ROUND(E{r}*{rate_of},0))')
    ws.cell(row=r, column=11, value=f'=IF(A{r}="","",ROUND(D{r}*{rate_of},0))')
style_block(ws, INS_R0, INS_R1, 11, {9, 10, 11},
            {4: YEN, 5: YEN, 6: YEN, 9: YEN, 10: YEN, 11: YEN}, example_row=INS_R0)
ws.cell(row=INS_R1 + 1, column=1, value="合計（見本行は数えません）").font = H2
for col in (10, 11):
    c = ws.cell(row=INS_R1 + 1, column=col,
                value=f"=SUM({get_column_letter(col)}{INS_R0 + 1}:{get_column_letter(col)}{INS_R1})")
    c.font, c.number_format, c.border = H2, YEN, BOX
dv3 = DataValidation(type="list", formula1='"JPY,USD,EUR"', allow_blank=True)
ws.add_data_validation(dv3); dv3.add(f"C{INS_R0}:C{INS_R1}")
ws.cell(row=6, column=4).comment = Comment(
    "これまでに払った保険料の合計です。取得原価にあたります。", "入力シート", width=260, height=70)
ws.cell(row=6, column=7).comment = Comment(
    "解約返戻金が「いつ時点」の数字かを入れます。基準日とズレていて構いません（画面にそのまま出ます）。",
    "入力シート", width=280, height=90)
note(ws, INS_R1 + 3, "※ 解約返戻金がまだ分からないときは、払込保険料累計と同じ額を入れ、"
                     "「補足」に「解約返戻金 未確認（暫定）」と書いてください。含み損益は ±0 になります。")

# ============================================================
# その他の借入
# ============================================================
ws = wb.create_sheet("その他の借入")
ws.sheet_view.showGridLines = False
ws["A1"] = "不動産以外の借入　（1行に1つのローン）"; ws["A1"].font = TITLE
note(ws, 2, "車のローン、教育ローン、多目的ローンなど。不動産のローンは「不動産」タブに書くので、ここには書きません。")
note(ws, 3, "借入がなければ、すべて空のままで構いません。")
head_row(ws, 5, ["借入の名前", "種類", "借入先", "当初借入額", "残債", "金利(年%)", "毎月の返済額", "返済済み"],
         [20, 12, 16, 15, 15, 11, 15, 14])
OL_R0, OL_R1 = 6, 15
for col, val in zip(range(1, 8), ["（例）自動車ローン", "車", "D銀行", 3000000, 1800000, 2.5, 45000]):
    ws.cell(row=OL_R0, column=col, value=val)
for r in range(OL_R0, OL_R1 + 1):
    ws.cell(row=r, column=8, value=f"=IF(A{r}=\"\",\"\",D{r}-E{r})")
style_block(ws, OL_R0, OL_R1, 8, {8}, {4: YEN, 5: YEN, 6: RATE, 7: YEN, 8: YEN}, example_row=OL_R0)

# ============================================================
# 検算
# ============================================================
ws = wb.create_sheet("検算")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 58
ws["A1"] = "検算　（自動計算。送る前にここを見てください）"; ws["A1"].font = TITLE
note(ws, 2, "この表はすべて自動です。画面に出る数字と同じものを先に確かめられます。")

def re_sum_ex(col):
    """見本行を除いた合計。見本行 RE_R0 は数えず、RE_R0+1 行目から数える。"""
    return f"SUM('不動産'!${col}${RE_R0 + 1}:${col}${RE_R1})"

ws["A4"] = "金融資産"; ws["A4"].font = H2
SEC_V = f"'有価証券'!H{SEC_R1 + 1}"          # 円換算した評価額 合計
SEC_C = f"'有価証券'!I{SEC_R1 + 1}"          # 円換算した取得原価 合計
INS_V = f"'保険'!J{INS_R1 + 1}"              # 円換算した解約返戻金 合計
INS_C = f"'保険'!K{INS_R1 + 1}"              # 円換算した払込保険料 合計
DEP_V = f"'預貯金'!D{DEP_R1 + 1}"

items = [
    ("有価証券（評価額・円）", f"={SEC_V}", YEN, "円換算した評価額の合計"),
    ("　うち含み損益",        f"={SEC_V}-{SEC_C}", YEN, "円換算した評価額 − 円換算した取得原価"),
    ("保険（解約返戻金・円）", f"={INS_V}", YEN, "死亡保険金額は含めません"),
    ("　うち含み損益",        f"={INS_V}-{INS_C}", YEN, "解約返戻金 − 払込保険料（保障の対価を含む）"),
    ("預貯金",                f"={DEP_V}", YEN, "円換算した残高の合計"),
    ("金融資産合計",          f"={SEC_V}+{INS_V}+{DEP_V}", YEN,
                              "有価証券 ＋ 保険 ＋ 預貯金（不動産は含めません）"),
    ("　うち含み損益",        f"={SEC_V}+{INS_V}-{SEC_C}-{INS_C}", YEN, "画面のサマリーに出る数字"),
]
r = 5
for label, formula, fmt, desc in items:
    ws.cell(row=r, column=1, value=label).font = H2 if "合計" in label else BODY
    c = ws.cell(row=r, column=2, value=formula)
    c.font, c.number_format, c.border, c.fill = (H2 if "合計" in label else CALC), fmt, BOX, FILL_CALC
    ws.cell(row=r, column=1).border = BOX
    ws.cell(row=r, column=3, value=desc).font = MUTED
    r += 1

r += 1
ws.cell(row=r, column=1, value="不動産").font = H2
r += 1
re_items = [
    ("総投資額",   f"={re_sum_ex('R')}", YEN, "自己資金 ＋ 当初借入額"),
    ("自己資金累計", f"={re_sum_ex('F')}", YEN, ""),
    ("借入総額",   f"={re_sum_ex('G')}", YEN, "当初の借入額の合計"),
    ("残債合計",   f"={re_sum_ex('H')}", YEN, ""),
    ("累計返済額", f"={re_sum_ex('S')}", YEN, "借入総額 − 残債合計"),
    ("返済進捗率", f"=IFERROR({re_sum_ex('S')}/{re_sum_ex('G')},0)", PCT1, "累計返済額 ÷ 借入総額"),
    ("月次収支の合計", f"={re_sum_ex('U')}", YEN, "家賃 − 返済 − 経費。マイナスなら持ち出しです"),
]
for label, formula, fmt, desc in re_items:
    ws.cell(row=r, column=1, value=label).font = BODY
    c = ws.cell(row=r, column=2, value=formula)
    c.font, c.number_format, c.border, c.fill = CALC, fmt, BOX, FILL_CALC
    ws.cell(row=r, column=1).border = BOX
    ws.cell(row=r, column=3, value=desc).font = MUTED
    r += 1

r += 1
ws.cell(row=r, column=1, value="負債").font = H2
r += 1
debt_items = [
    ("不動産ローン 残債", f"={re_sum_ex('H')}", YEN, ""),
    ("その他の借入 残債", f"=SUM('その他の借入'!E{OL_R0 + 1}:E{OL_R1})", YEN, "車・多目的など"),
    ("総負債", f"={re_sum_ex('H')}+SUM('その他の借入'!E{OL_R0 + 1}:E{OL_R1})", YEN, "不動産 ＋ その他"),
]
for label, formula, fmt, desc in debt_items:
    ws.cell(row=r, column=1, value=label).font = H2 if label == "総負債" else BODY
    c = ws.cell(row=r, column=2, value=formula)
    c.font, c.number_format, c.border, c.fill = (H2 if label == "総負債" else CALC), fmt, BOX, FILL_CALC
    ws.cell(row=r, column=1).border = BOX
    ws.cell(row=r, column=3, value=desc).font = MUTED
    r += 1

r += 2
ws.cell(row=r, column=1, value="つじつまの確認").font = H2
note(ws, r + 1, "「要確認」が出たら、その行の説明どおりに直してください。")
r += 2
checks = [
    ("残債が当初借入額より大きい物件",
     f'=IF(SUMPRODUCT((\'不動産\'!A{RE_R0+1}:A{RE_R1}<>"")*(\'不動産\'!H{RE_R0+1}:H{RE_R1}>\'不動産\'!G{RE_R0+1}:G{RE_R1}))>0,"要確認","OK")',
     "残債は当初借入額以下になるはずです"),
    ("残債が当初借入額より大きい借入",
     f'=IF(SUMPRODUCT((\'その他の借入\'!A{OL_R0+1}:A{OL_R1}<>"")*(\'その他の借入\'!E{OL_R0+1}:E{OL_R1}>\'その他の借入\'!D{OL_R0+1}:D{OL_R1}))>0,"要確認","OK")',
     "同上"),
    ("マイナスの金額が入っている",
     f'=IF(SUMPRODUCT((\'不動産\'!F{RE_R0+1}:H{RE_R1}<0)*1)+SUMPRODUCT((\'有価証券\'!D{SEC_R0}:E{SEC_R1}<0)*1)>0,"要確認","OK")',
     "金額はすべて 0 以上で入れます"),
    ("為替レートの無い通貨がある",
     f'=IF(SUMPRODUCT(ISERROR(MATCH(\'有価証券\'!C{SEC_R0}:C{SEC_R1},\'基本情報\'!$A$10:$A$13,0))*1)>0,"要確認","OK")',
     "「基本情報」タブの為替レートに、使った通貨を足してください"),
    ("氏名と基準日が入っている",
     '=IF(OR(\'基本情報\'!B4="",\'基本情報\'!B5=""),"要確認","OK")',
     "どちらも必須です"),
    ("借入期間の「か月」が0〜11か",
     f'=IF(SUMPRODUCT((\'不動産\'!A{RE_R0+1}:A{RE_R1}<>"")*((\'不動産\'!L{RE_R0+1}:L{RE_R1}<0)+'
     f'(\'不動産\'!L{RE_R0+1}:L{RE_R1}>11)))>0,"要確認","OK")',
     "12か月以上は「年」の側に繰り上げてください"),
    ("保険にレートの無い通貨がある",
     f'=IF(SUMPRODUCT((\'保険\'!A{INS_R0+1}:A{INS_R1}<>"")*'
     f'ISERROR(MATCH(\'保険\'!C{INS_R0+1}:C{INS_R1},\'基本情報\'!$A$10:$A$13,0)))>0,"要確認","OK")',
     "「基本情報」タブの為替レートに、使った通貨を足してください"),
]
for label, formula, desc in checks:
    ws.cell(row=r, column=1, value=label).font = BODY
    c = ws.cell(row=r, column=2, value=formula)
    c.font, c.border, c.fill = CALC, BOX, FILL_CALC
    c.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=1).border = BOX
    ws.cell(row=r, column=3, value=desc).font = MUTED
    r += 1

from openpyxl.formatting.rule import CellIsRule
ws.conditional_formatting.add(
    f"B{r - len(checks)}:B{r - 1}",
    CellIsRule(operator="equal", formula=['"要確認"'], fill=FILL_NG,
               font=Font(name=F, size=10, bold=True, color="9C0006")))

# 印刷しても横で切れないように、各シートを横向き・幅1ページに収める
for sh in wb.worksheets:
    sh.page_setup.orientation = "landscape"
    sh.page_setup.fitToWidth = 1
    sh.page_setup.fitToHeight = 0
    sh.sheet_properties.pageSetUpPr.fitToPage = True

wb.save("data/input-sheet.xlsx")
print("wrote data/input-sheet.xlsx")
